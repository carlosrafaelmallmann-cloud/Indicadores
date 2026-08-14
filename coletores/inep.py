"""Coletor do IDEB — Instituto Nacional de Estudos e Pesquisas Educacionais.

O INEP não expõe API. As planilhas de divulgação (xlsx) trazem, cada uma, a
série histórica completa em colunas — uma edição contém todos os anos desde
2005. A cada divulgação (bienal), basta depositar os arquivos em dados/inep/
e a coleta seguinte incorpora tudo.

Arquivos esperados em dados/inep/ (xlsx ou zip contendo xlsx):
    divulgacao_anos_iniciais_municipios_AAAA.xlsx
    divulgacao_anos_finais_municipios_AAAA.xlsx
    divulgacao_ensino_medio_municipios_AAAA.xlsx
    divulgacao_anos_iniciais_escolas_AAAA.xlsx
    divulgacao_anos_finais_escolas_AAAA.xlsx
    divulgacao_ensino_medio_escolas_AAAA.xlsx
    divulgacao_*_brasil_estados_AAAA.xlsx   (para a comparação com o RS)

O que for extraído é gravado em dados/ideb.csv, que passa a ser a fonte
canônica: mesmo sem as planilhas, o painel continua funcionando a partir do
CSV versionado, que também pode ser editado à mão.

Estrutura de dados/ideb.csv (formato longo):
    etapa,ambito,rede,nome,codigo,ano,ideb,meta
"""
from __future__ import annotations

import csv
import logging
import re
import zipfile
from pathlib import Path

from .comum import (COD_IBGE, DIR_DADOS, Indicador, ResultadoColeta,
                    normalizar, para_numero)

log = logging.getLogger("inep")

DIR_PLANILHAS = DIR_DADOS / "inep"
ARQ_IDEB = DIR_DADOS / "ideb.csv"
COLUNAS = ["etapa", "ambito", "rede", "nome", "codigo", "ano", "ideb", "meta"]

URL_FONTE = ("https://www.gov.br/inep/pt-br/areas-de-atuacao/"
             "pesquisas-estatisticas-e-indicadores/ideb/resultados")

ETAPAS = {
    "anos_iniciais": "Anos iniciais",
    "anos iniciais": "Anos iniciais",
    "anos_finais": "Anos finais",
    "anos finais": "Anos finais",
    "ensino_medio": "Ensino médio",
    "ensino medio": "Ensino médio",
}
ORDEM_ETAPAS = ["Anos iniciais", "Anos finais", "Ensino médio"]


# ---------------------------------------------------------------------------
# Leitura das planilhas do INEP
# ---------------------------------------------------------------------------
def _identificar_etapa(nome_arquivo: str) -> str | None:
    rotulo = normalizar(nome_arquivo).replace("-", "_")
    for chave, etapa in ETAPAS.items():
        if normalizar(chave).replace(" ", "_") in rotulo.replace(" ", "_"):
            return etapa
    return None


def _identificar_ambito(nome_arquivo: str) -> str:
    rotulo = normalizar(nome_arquivo)
    if "escola" in rotulo:
        return "escola"
    if "municipio" in rotulo:
        return "municipio"
    return "uf"


def _preencher(linha: list) -> list:
    """Células mescladas vêm vazias fora da âncora: repete o último rótulo."""
    saida, ultimo = [], ""
    for celula in linha:
        texto = "" if celula is None else str(celula).strip()
        if texto:
            ultimo = texto
        saida.append(ultimo)
    return saida


def _mapear_cabecalho(linhas: list[list]) -> dict | None:
    """Localiza a linha de anos e classifica cada coluna por bloco (IDEB, meta)
    e por identificação (código, nome, rede)."""
    for indice, linha in enumerate(linhas):
        anos = [(pos, int(str(v)[:4])) for pos, v in enumerate(linha)
                if v is not None and re.fullmatch(r"\s*20[0-2]\d(\.0)?\s*", str(v).strip())]
        if len(anos) < 4:
            continue

        blocos = _preencher(linhas[indice - 1]) if indice else [""] * len(linha)
        if all(not b for b in blocos) and indice >= 2:
            blocos = _preencher(linhas[indice - 2])

        ideb, metas = {}, {}
        for pos, ano in anos:
            rotulo = normalizar(blocos[pos] if pos < len(blocos) else "")
            if "projec" in rotulo or "meta" in rotulo:
                metas[ano] = pos
            elif "ideb" in rotulo:
                ideb[ano] = pos

        if not ideb:   # planilha só do IDEB, sem rótulo de bloco acima
            ideb = {ano: pos for pos, ano in anos}

        identificacao: dict[str, int] = {}
        for busca in linhas[:indice + 1]:
            for pos, valor in enumerate(busca):
                rotulo = normalizar(valor)
                if not rotulo:
                    continue
                if "escola" in rotulo and "codigo" in rotulo:
                    identificacao.setdefault("codigo_escola", pos)
                elif "escola" in rotulo and "nome" in rotulo:
                    identificacao.setdefault("nome_escola", pos)
                elif "municipio" in rotulo and "codigo" in rotulo:
                    identificacao.setdefault("codigo_municipio", pos)
                elif "municipio" in rotulo and "nome" in rotulo:
                    identificacao.setdefault("nome_municipio", pos)
                elif rotulo.startswith("rede") or "dependencia" in rotulo:
                    identificacao.setdefault("rede", pos)
                elif rotulo in {"uf", "sigla da uf", "estado"} or "sigla" in rotulo:
                    identificacao.setdefault("uf", pos)

        return {"linha_anos": indice, "ideb": ideb, "metas": metas,
                "identificacao": identificacao}
    return None


def _ler_planilha(caminho: Path, etapa: str, ambito: str) -> list[dict]:
    from openpyxl import load_workbook

    livro = load_workbook(caminho, read_only=True, data_only=True)
    pagina = livro[livro.sheetnames[0]]
    linhas = [list(l) for l in pagina.iter_rows(max_row=20, values_only=True)]

    mapa = _mapear_cabecalho(linhas)
    if mapa is None:
        log.error("não identifiquei o cabeçalho de %s — primeiras linhas: %s",
                  caminho.name, [l[:6] for l in linhas[:6]])
        livro.close()
        return []

    ident = mapa["identificacao"]
    registros: list[dict] = []
    interessa_uf = ambito == "uf"

    for linha in pagina.iter_rows(min_row=mapa["linha_anos"] + 2, values_only=True):
        if linha is None or all(c is None for c in linha):
            continue

        def celula(chave: str) -> str:
            pos = ident.get(chave)
            if pos is None or pos >= len(linha) or linha[pos] is None:
                return ""
            return str(linha[pos]).strip()

        codigo_municipio = re.sub(r"\D", "", celula("codigo_municipio"))
        if interessa_uf:
            rotulo = normalizar(celula("nome_municipio") or celula("uf"))
            if rotulo not in {"rs", "rio grande do sul", "brasil"}:
                continue
            nome = "Brasil" if rotulo == "brasil" else "Rio Grande do Sul"
            codigo = rotulo
        else:
            if codigo_municipio != COD_IBGE:
                continue
            if ambito == "escola":
                nome = celula("nome_escola") or "Escola sem identificação"
                codigo = re.sub(r"\D", "", celula("codigo_escola"))
            else:
                nome = "Estrela"
                codigo = COD_IBGE

        rede = celula("rede") or "Total"

        for ano, pos in sorted(mapa["ideb"].items()):
            valor = para_numero(linha[pos]) if pos < len(linha) else None
            meta_pos = mapa["metas"].get(ano)
            meta = para_numero(linha[meta_pos]) if meta_pos is not None and meta_pos < len(linha) else None
            if valor is None and meta is None:
                continue
            registros.append({"etapa": etapa, "ambito": ambito, "rede": rede,
                              "nome": nome, "codigo": codigo, "ano": str(ano),
                              "ideb": valor, "meta": meta})

    livro.close()
    log.info("%s: %d registros (%s / %s)", caminho.name, len(registros), etapa, ambito)
    return registros


def _extrair_planilhas() -> list[dict]:
    if not DIR_PLANILHAS.exists():
        return []
    registros: list[dict] = []
    for caminho in sorted(DIR_PLANILHAS.iterdir()):
        if caminho.suffix.lower() == ".zip":
            with zipfile.ZipFile(caminho) as pacote:
                for interno in pacote.namelist():
                    if not interno.lower().endswith(".xlsx"):
                        continue
                    etapa = _identificar_etapa(caminho.name) or _identificar_etapa(interno)
                    if etapa is None:
                        continue
                    destino = DIR_PLANILHAS / f"__tmp_{Path(interno).name}"
                    destino.write_bytes(pacote.read(interno))
                    try:
                        registros += _ler_planilha(
                            destino, etapa,
                            _identificar_ambito(caminho.name + interno))
                    finally:
                        destino.unlink(missing_ok=True)
        elif caminho.suffix.lower() == ".xlsx" and not caminho.name.startswith("__tmp_"):
            etapa = _identificar_etapa(caminho.name)
            if etapa is None:
                log.warning("etapa não identificada pelo nome de %s — ignorado", caminho.name)
                continue
            registros += _ler_planilha(caminho, etapa, _identificar_ambito(caminho.name))
    return registros


# ---------------------------------------------------------------------------
# CSV canônico
# ---------------------------------------------------------------------------
def _ler_csv() -> list[dict]:
    if not ARQ_IDEB.exists():
        return []
    with ARQ_IDEB.open(encoding="utf-8-sig", newline="") as arquivo:
        registros = []
        for linha in csv.DictReader(arquivo):
            linha["ideb"] = para_numero(linha.get("ideb"))
            linha["meta"] = para_numero(linha.get("meta"))
            if linha.get("etapa") and linha.get("ano"):
                registros.append(linha)
        return registros


def _gravar_csv(registros: list[dict]) -> None:
    ARQ_IDEB.parent.mkdir(parents=True, exist_ok=True)
    with ARQ_IDEB.open("w", encoding="utf-8", newline="") as arquivo:
        escritor = csv.DictWriter(arquivo, fieldnames=COLUNAS, extrasaction="ignore")
        escritor.writeheader()
        chave = lambda r: (ORDEM_ETAPAS.index(r["etapa"]) if r["etapa"] in ORDEM_ETAPAS else 9,
                           r["ambito"], r["nome"], r["rede"], r["ano"])
        for registro in sorted(registros, key=chave):
            escritor.writerow(registro)


def _mesclar(base: list[dict], novos: list[dict]) -> list[dict]:
    """Registros das planilhas prevalecem sobre os do CSV."""
    indexado = {(r["etapa"], r["ambito"], r["nome"], r["rede"], r["ano"]): r for r in base}
    for registro in novos:
        indexado[(registro["etapa"], registro["ambito"],
                  registro["nome"], registro["rede"], registro["ano"])] = registro
    return list(indexado.values())


# ---------------------------------------------------------------------------
# Consolidação do painel
# ---------------------------------------------------------------------------
def _rotulo_serie(registro: dict) -> str | None:
    rede = normalizar(registro["rede"])
    if registro["ambito"] == "uf":
        return registro["nome"]
    if registro["ambito"] == "municipio":
        if "municipal" in rede:
            return "Estrela — rede municipal"
        if "estadual" in rede:
            return "Estrela — rede estadual"
        if "publica" in rede or rede in {"total", ""}:
            return "Estrela — rede pública"
    return None


CORES = {
    "Estrela — rede municipal": "#157A72",
    "Estrela — rede estadual": "#B8892B",
    "Estrela — rede pública": "#26414F",
    "Rio Grande do Sul": "#9E3B2E",
    "Brasil": "#6C7C85",
}


def _consolidar(registros: list[dict]) -> dict:
    etapas = []
    for etapa in ORDEM_ETAPAS:
        deste = [r for r in registros if r["etapa"] == etapa]
        if not deste:
            continue
        anos = sorted({r["ano"] for r in deste})

        agrupado: dict[str, dict[str, float]] = {}
        for registro in deste:
            rotulo = _rotulo_serie(registro)
            if rotulo and registro["ideb"] is not None:
                agrupado.setdefault(rotulo, {})[registro["ano"]] = registro["ideb"]

        series = [{"rotulo": rotulo, "cor": CORES.get(rotulo, "#26414F"),
                   "valores": [valores.get(ano) for ano in anos]}
                  for rotulo, valores in agrupado.items()
                  if any(v is not None for v in valores.values())]
        series.sort(key=lambda s: list(CORES).index(s["rotulo"])
                    if s["rotulo"] in CORES else 9)

        escolas: dict[str, dict[str, float]] = {}
        for registro in deste:
            if registro["ambito"] == "escola" and registro["ideb"] is not None:
                escolas.setdefault(registro["nome"], {})[registro["ano"]] = registro["ideb"]

        etapas.append({
            "etapa": etapa,
            "anos": anos,
            "series": series,
            "escolas": [{"nome": nome, "valores": [valores.get(ano) for ano in anos]}
                        for nome, valores in sorted(escolas.items())],
        })
    return {"etapas": etapas}


def coletar() -> ResultadoColeta:
    base = _ler_csv()
    try:
        novos = _extrair_planilhas()
    except Exception as erro:  # noqa: BLE001
        log.error("falha ao ler as planilhas do INEP: %s", erro)
        novos = []

    registros = _mesclar(base, novos) if novos else base
    if novos:
        _gravar_csv(registros)

    if not registros:
        return ResultadoColeta(
            fonte="INEP / IDEB", status="falha",
            mensagem="sem dados: deposite as planilhas em dados/inep/ ou "
                     "preencha dados/ideb.csv")

    painel = _consolidar(registros)
    indicadores: list[Indicador] = []

    for bloco in painel["etapas"]:
        principal = next((s for s in bloco["series"]
                          if s["rotulo"].startswith("Estrela")), None)
        if principal is None:
            continue
        pares = [(ano, valor) for ano, valor in zip(bloco["anos"], principal["valores"])
                 if valor is not None]
        if not pares:
            continue
        ano, valor = pares[-1]
        indicadores.append(Indicador(
            id=f"ideb_{normalizar(bloco['etapa']).replace(' ', '_')}",
            categoria="Educação",
            indicador=f"IDEB — {bloco['etapa'].lower()}",
            valor=valor,
            unidade="pontos",
            periodo=ano,
            fonte="INEP",
            fonte_detalhe=principal["rotulo"],
            url=URL_FONTE,
            formato="numero",
            serie=[{"periodo": a, "valor": v} for a, v in pares],
            observacao="Índice de Desenvolvimento da Educação Básica, apurado "
                       "bienalmente pelo INEP em escala de 0 a 10.",
        ))

    total_escolas = sum(len(b["escolas"]) for b in painel["etapas"])
    mensagem = (f"{len(painel['etapas'])} etapas; {total_escolas} séries por escola"
                + (f"; {len(novos)} registros lidos das planilhas" if novos
                   else "; nenhuma planilha nova em dados/inep/"))

    resultado = ResultadoColeta(fonte="INEP / IDEB", status="ok",
                                mensagem=mensagem, indicadores=indicadores)
    resultado.painel_educacao = painel  # type: ignore[attr-defined]
    return resultado
